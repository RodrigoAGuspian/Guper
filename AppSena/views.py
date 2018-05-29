from django.shortcuts import render, redirect
from .models import *
from .forms import * 
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate

from openpyxl import load_workbook

# VIEWS
#=======================INDEX=====================#
def view_index(request):
	if request.user.is_authenticated:
		return render(request, 'index.html')
	else:
		return redirect('url_login')
#=================================================#

#=======================FICHA=====================#
def view_lista_fichas(request):
	ficha = Ficha.objects.all()
	return render(request, 'lista_fichas.html', locals())
#=================================================#

#=================AGREGAR FICHA===================#
def view_agregar_ficha(request):
	if request.method == "POST":
		formulario_ficha = agregar_ficha_form(request.POST)
		if formulario_ficha.is_valid():
			ficha = formulario_ficha.save(commit = False)
			ficha.save()
			formulario_ficha.save_m2m()
			return redirect('url_lista_fichas')
	else:
		formulario_ficha = agregar_ficha_form()
	return render(request, 'agregar_ficha.html', locals())
#==================================================#

#================AGREGAR PERSONA FICHA=============#
def view_agregar_persona_ficha(request):
	if request.method == "POST":
		formulario_persona_ficha = agregar_persona_ficha_form(request.POST)
		if formulario_persona_ficha.is_valid():
			persona_ficha = formulario_persona_ficha.save(commit = False)
			persona_ficha.save()
			formulario_persona_ficha.save()
			return redirect('url_lista_fichas')
	else:
		formulario_persona_ficha = agregar_persona_ficha_form()
	return render(request, 'agregar_persona_ficha.html', locals())
#===================================================#

#===================EDITAR FICHA====================#
def view_editar_ficha(request, id_ficha):
	ficha = Ficha.objects.get(id = id_ficha)
	if request.method == "POST":
		formulario = agregar_ficha_form(request.POST, instance = ficha)
		if formulario.is_valid():
			ficha = formulario.save()
			return redirect('url_lista_fichas')
	else:
		formulario = agregar_ficha_form(instance = ficha)
	return render(request, 'editar_ficha.html', locals())
#===================================================#

#==================BORRAR FICHA=====================#
def view_eliminar_ficha(request, id_ficha):
	ficha = Ficha.objects.get(id = id_ficha)
	ficha.delete()
	return redirect('url_lista_fichas')
#=================================================#

#=====================PERMISO=====================#
def view_lista_permisos(request):
	permiso = Permiso.objects.all()
	return render(request, 'lista_permisos.html', locals())
#=================================================#

#=====================PERSONA=====================#

#=================REGISTRAR INSTRUCTOR================#

#========== REGISTRAR ADMINISTRADOR =============#
def view_registrar_administrador(request):
	rtemp = Rol.objects.filter(rol='ADMIN')
	roltemp = ""
	for i in rtemp:
		roltemp = i.rol
	if roltemp == "":
		r = Rol.objects.create(rol="ADMIN")
		r.save()
	else:
		rl = Rol.objects.get(rol = "ADMIN")
		if request.method == 'POST':
			form = agregar_persona_form(request.POST, request.FILES)
			form2 = agregar_user_form(request.POST, request.FILES)
			if form.is_valid() and form2.is_valid():

				nom1 = form.cleaned_data['primerNombre']
				nom2 = form.cleaned_data['segundoNombre']
				ape1 = form.cleaned_data['primerApellido']
				ape2 = form.cleaned_data['segundoApellido']
				cont = form.cleaned_data['contacto']

				email = form2.cleaned_data['username']
				documento = form2.cleaned_data['password']
		
				u = User.objects.create_user(username=email,password=documento, email=email)
				u.save()

				persona = Persona.objects.create(documentoIdentidad=documento ,primerNombre=nom1 ,segundoNombre=nom2 ,primerApellido=ape1 ,segundoApellido=ape2 ,contacto=cont ,usuario=u)
				persona.save()


				aprendiz = Rol_persona.objects.create(rol=rl ,persona=persona)
				aprendiz.save()

				return redirect ('url_lista_aprendices')

		else:
			form = agregar_persona_form()
			form2 = agregar_user_form()
	return render (request, 'agregar_administrador.html', locals())
#=================================================#

def view_agregar_instructor(request):
	ins=Rol.objects.get(rol='Instructor')
	if request.method == 'POST':
		formulario = agregar_persona_form(request.POST, request.FILES)
		formulario2= agregar_user_form(request.POST, request.FILES)
		if formulario.is_valid() and formulario2.is_valid():

			nom =formulario.cleaned_data['nombres']
			ape =formulario.cleaned_data['apellidos']
			tel =formulario.cleaned_data['telefono']

			email =formulario2.cleaned_data['username']
			documento =formulario2.cleaned_data['password']

			u = User.objects.create_user(username=email, password=documento)
			u.save()

			persona = Persona.objects.create(documentoIdentidad=documento,nombres=nom,apellidos=ape,telefono=tel,usuario=u)
			persona.save()

			vigilante = Rol_persona.objects.create(rol=ins, persona=persona)
			vigilante.save()
			
			return redirect ('url_index')

	else:
		formulario=agregar_persona_form()
		formulario2= agregar_user_form()

	return render(request, 'agregar_instructor.html', locals())

#=======================REGISTRAR VIGILANTE===============================
def view_agregar_vigilante(request):
	vi=Rol.objects.get(rol='Vigilante')
	if request.method == 'POST':
		formulario = agregar_persona_form(request.POST, request.FILES)
		formulario2= agregar_user_form(request.POST, request.FILES)
		if formulario.is_valid() and formulario2.is_valid():

			nom =formulario.cleaned_data['nombres']
			ape =formulario.cleaned_data['apellidos']
			tel =formulario.cleaned_data['telefono']

			email =formulario2.cleaned_data['username']
			documento =formulario2.cleaned_data['password']

			u = User.objects.create_user(username=email, password=documento)
			u.save()

			persona = Persona.objects.create(documentoIdentidad=documento,nombres=nom,apellidos=ape,telefono=tel,usuario=u)
			persona.save()

			vigilante = Rol_persona.objects.create(rol=vi, persona=persona)
			vigilante.save()
			
			return redirect ('url_index')

	else:
		formulario=agregar_persona_form()
		formulario2= agregar_user_form()

	return render(request, 'agregar_instructor.html', locals())	
#=================================================#

#==============LISTA_APRENDICES====================#
def view_lista_aprendices(request):
	r = Rol.objects.get(rol='Aprendiz')
	rp = Rol_persona.objects.filter(rol=r)
	#aprendices=Persona.objects.filter()
	return render (request, 'lista_aprendices.html', locals())
#=================================================#

#==============AGREGAR_APRENDIZ===================#
def view_agregar_aprendiz(request):

	rl = Rol.objects.get(rol='aprendiz')
	if request.method == 'POST':
		formulario = agregar_persona_form(request.POST, request.FILES)
		formulario2 = agregar_user_form(request.POST, request.FILES)
		if formulario.is_valid() and formulario2.is_valid():

			nom =formulario.cleaned_data['nombres']
			ape =formulario.cleaned_data['apellidos']
			tel =formulario.cleaned_data['telefono']

			email =formulario2.cleaned_data['username']
			documento =formulario2.cleaned_data['password']

			u = User.objects.create_user(username=email, password=documento)
			u.save()

			persona = Persona.objects.create(documentoIdentidad=documento,nombres=nom,apellidos=ape,telefono=tel,usuario=u)
			persona.save()


			aprendiz = Rol_persona.objects.create(rol=rl ,persona=persona)
			aprendiz.save()

			return redirect ('url_lista_aprendices')

	else:
		formulario = agregar_persona_form()
		formulario2 = agregar_user_form()
	return render (request, 'agregar_aprendiz.html', locals())
#=================================================#

#==============EDITAR_APRENDIZ====================#
def view_editar_aprendiz(request, id_aprendiz):

	apr = Persona.objects.get(id=id_aprendiz)

	if request.method == 'POST':
		formulario = agregar_persona_form(request.POST, request.FILES, instance=apr)
		formulario2 = editar_user_form(request.POST, request.FILES, instance=apr.usuario)
		if formulario.is_valid() and formulario2.is_valid():

			apr = formulario.save()
			usu = formulario2.save()

			return redirect ('url_lista_aprendices')
	else:
		formulario = agregar_persona_form(instance=apr)
		formulario2 = editar_user_form(instance=apr.usuario)
	return render (request, 'agregar_aprendiz.html', locals())
#=================================================#

#==============ELIMINAR_APRENDIZ====================#
def view_eliminar_aprendiz(request, id_aprendiz):
	aprendiz = Persona.objects.get(id=id_aprendiz)
	id_user = aprendiz.usuario.id
	user = User.objects.get(id=id_user)
	aprendiz.delete()
	user.delete()
	return redirect ('url_lista_aprendices')
#=================================================#

#=========VIEW_AGREGAR_DESDE_EXCEL=================#
def view_agregar_aprendiz_excel(request):
	fichas = Ficha.objects.all()
	if request.method == 'POST':
		form = cargar_excel_form(request.POST, request.FILES)
		if form.is_valid():
			excel = form.cleaned_data['excel']
			return redirect('url_lista_aprendices')#'url_registros_excel')
	else:
		form = cargar_excel_form()
	return render(request, 'agregar_aprendices_excel.html', locals())
#=================================================#

#=============AGREGAR_DESDE_EXCEL=================#
def cargar_excel(request):
	rl = Rol.objects.get(rol='Aprendiz')
	FILE_PATH = 'media/registros.xlsx' #define el achivo excel en una variable
	SHEET = 'Hoja1'                  #define la hoja de excel en una variable

	wb = load_workbook(FILE_PATH, read_only=True)# cargo el archivo en una variable
	sheet = wb[SHEET]                            # cargo la hoja en una variable
	
	for row in sheet.iter_rows(min_row=2):						
		documento = row[0].value
		nom = row[1].value
		ape = row[2].value
		email = row[3].value
		#ape2 = row[3].value
		tel = row[5].value

		u = User.objects.create_user(username=email,password=documento, email=email)
		u.save()
		
		persona = Persona.objects.create(documentoIdentidad=documento,nombres=nom,apellidos=ape,telefono=tel,usuario=u)
		persona.save()


		aprendiz = Rol_persona.objects.create(rol=rl ,persona=persona)
		aprendiz.save()
	lista = sheet.iter_rows(min_row=2)

	return render(request, 'registros_excel.html', locals())
#==================================================


#====================PROGRAMA=====================#
def view_lista_programas(request):
	#LISTAR
	programa = Programa.objects.all()
	return render(request, 'lista_programas.html', locals())

#===============AGREGAR PROGRAMA==============#
def view_agregar_programa(request):	
	if request.method == 'POST':
		formulario = agregar_programa_form(request.POST, request.FILES)
		if formulario.is_valid():
			pro = formulario.save(commit=False)
			pro.save()
			return redirect('url_lista_programas')
	else:
		formulario = agregar_programa_form()

	return render (request, 'agregar_programa.html',locals())

#===================EDITAR PROGRAMA==============#	
def view_editar_programa(request, id_programa):
	pro = Programa.objects.get(id=id_programa)
	if request.method == "POST":
		formulario = agregar_programa_form(request.POST, request.FILES, instance=pro)
		if formulario.is_valid():
			pro = formulario.save()
			return redirect('url_lista_programas')
	else:
		formulario = agregar_programa_form(instance= pro)
	return render(request, 'agregar_programa.html', locals())

#===================ELIMINAR PROGRAMA==============#
def view_eliminar_programa(request, id_programa):
	pro = Programa.objects.get(id=id_programa)
	pro.delete()
	return redirect('url_lista_programas')
	
#=================================================#

#=====================USUARIO=====================#
def view_usuario(request):
	usuario = User.objects.all()
	return render(request, 'usuario/index.html', locals())
#=================================================#

#===================VER_ROL=======================#
def view_lista_roles(request):
	rol = Rol.objects.all()
	return render(request,'lista_roles.html',locals())

#===================AGREGAR ROL===================#
def view_agregar_rol(request):
	rol = Rol.objects.all()
	if request.method== 'POST':
		formulario= agregar_rol_form(request.POST,request.FILES)
		if formulario.is_valid():
			rol= formulario.save(commit= False)
			rol.save()
			formulario.save_m2m()
			return redirect('url_lista_roles')

	else:
		formulario = agregar_rol_form()
	return render(request,'agregar_rol.html',locals())


#===================EDITAR_ROL=====================#
def view_editar_rol(request, id_rol):
	rol= Rol.objects.get(id=id_rol)
	if request.method== 'POST':
		formulario= agregar_rol_form(request.POST, request.FILES, instance=rol)
		if formulario.is_valid():
			rol= formulario.save()
			return redirect('url_lista_roles')
	else:
		formulario= agregar_rol_form(instance= rol)
	return render(request, 'agregar_rol.html',locals())

#===================ELIMINAR_ROL===================#
def view_eliminar_rol(request, id_rol):
	rol= Rol.objects.get(id= id_rol)
	rol.delete()
	return redirect('url_lista_roles')

#==================== LOGIN ======================#
#@login_required  DECORADORES
def view_login(request):
	if request.user.is_authenticated:
		return redirect('url_index')
	else:
		loginf = login_form()
		if request.method == "POST":
			loginf = login_form(request.POST)
			if loginf.is_valid():
				usuario = loginf.cleaned_data['usuario']
				clave = loginf.cleaned_data["contraseña"]
				user = authenticate(username = usuario, password = clave)
				if user is not None and user.is_active:
					login(request, user)
					usuarioActivo = True
					return redirect('url_index')
				else:
					msj = "usuario o clave incorrecto"
		loginf = login_form()
		return render(request, "login.html", locals())
#=================================================#

#==================== logout ======================#
def view_logout(request):
	logout(request)
	return redirect('url_index')
#=================================================#

#=================================================#
