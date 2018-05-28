from django.shortcuts import render, redirect
from .models import *
from .forms import * 
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from datetime import datetime, date
from openpyxl import load_workbook

#=======================INDEX=====================#
def view_index(request):
	if request.user.is_authenticated:
		return render(request, 'index.html')
	else:
		return redirect('url_login')
#=================================================#

#=======================FICHA=====================#
def view_lista_fichas(request):
	ficha = Ficha.objects.all()
	if not ficha:
			msj = 'Lista Vacia. Click aqui para agregar'
	return render(request, 'lista_fichas.html', locals())
#=================================================#

#=================AGREGAR FICHA===================#
def view_agregar_ficha(request):

	msj= 'Agregar Ficha'

	if request.method == "POST":
		formulario_ficha = agregar_ficha_form(request.POST)
		if formulario_ficha.is_valid():

			h = date.today()
			fic = formulario_ficha.save(commit = False)

			fecha = formulario_ficha.cleaned_data['fechaFinEtapaLectiva']
			ambte = formulario_ficha.cleaned_data['ambiente']
			jrnad = formulario_ficha.cleaned_data['jornada']

			dispone = Ficha.objects.filter(ambiente=ambte, jornada=jrnad)

			if dispone:
				for i in dispone:
					ocupado = str(i.numeroFicha)+' de '+str(i.programa.nombre)
					msj = 'A ocurrido un error al intentar Agregar la Ficha'
					msj2='El anbiete seleccionado esta siendo ocupado en el mismo horario por la ficha numero: '+ocupado

			elif fecha>h:
				fic.save()
				return redirect('url_lista_fichas')
			else:
				msj = 'A ocurrido un error al intentar Agregar la Ficha'
				msj2='La Fecha Fin de Etapa Lectiva no puede ser anterior o igual a hoy'

	else:
		formulario_ficha = agregar_ficha_form()
	return render(request, 'agregar_ficha.html', locals())
#=================================================#

#================AGREGAR PERSONA FICHA============#
def view_agregar_persona_ficha(request):
	if request.method == "POST":
		formulario_persona_ficha = agregar_persona_ficha_form(request.POST)
		if formulario_persona_ficha.is_valid():
			persona_ficha = formulario_persona_ficha.save(commit = False)
			persona_ficha.save()
			formulario_persona_ficha.save()

			return redirect('url_lista_fichas')
	else:
		formulario_persona_ficha = agregar_persona_ficha_form()
	return render(request, 'agregar_persona_ficha.html', locals())
#=================================================#

#===================EDITAR FICHA==================#
def view_editar_ficha(request, id_ficha):

	msj='Editar Ficha'

	ficha = Ficha.objects.get(id = id_ficha)
	if request.method == "POST":
		formulario_ficha = agregar_ficha_form(request.POST, instance = ficha)
		if formulario_ficha.is_valid():

			h = date.today()

			fecha = formulario_ficha.cleaned_data['fechaFinEtapaLectiva']
			ambte = formulario_ficha.cleaned_data['ambiente']
			jrnad = formulario_ficha.cleaned_data['jornada']

			dispone = Ficha.objects.filter(ambiente=ambte, jornada=jrnad)
			print('-------------------------',dispone)

			if not dispone: #si el Query esta vacio el ambiente esta disponible
				print('>>>>>>>>>>Vacio')
				if fecha>h:
					formulario_ficha.save()
					return redirect('url_lista_fichas')
				else:
					msj = 'A ocurrido un error al intentar Editar la Ficha'
					msj2='La Fecha Fin de Etapa Lectiva no puede ser anterior o igual a hoy'
					print(msj2)
					#borrar
			else:
				for i in dispone:
					#VALORES SIN APLICAR CABIOS
					idfichDis =i.id  		       #id de la ficha SIN cambios
					numfichDis =i.numeroFicha      #numero de la ficha SIN cambios
					profichDis = i.programa.nombre #programa de la ficha SIN cambios
					ambienDis = i.ambiente
					jornaDis = i.jornada
					#VALORES CON LOS CABIOS QUE SE QUIEREN APLICAR
					ficid = ficha.id    		   #id de la ficha CON cambios
					ficNum = ficha.numeroFicha
					ficPro = ficha.programa.nombre
					ficAmb = ficha.ambiente
					FicJor = ficha.jornada

					print('====>>>> Sin cambios',idfichDis,numfichDis,profichDis,ambienDis,jornaDis)

					print('====>>>> Con cambios',ficid,ficNum,ficPro,ficAmb,FicJor)
					
					if idfichDis==ficid:
						print('----------->>>>Sin cambios en ambiente y/o jornada')
						if fecha>h:
							formulario_ficha.save()
							return redirect('url_lista_fichas')
						else:
							msj = 'A ocurrido un erro al intentar Editar la Ficha'
							msj2='La Fecha Fin de Etapa Lectiva no puede ser anterior o igual a hoy'
							print(msj2)
							#Borrar
					else:
						ocupado = numfichDis+' de '+profichDis
						msj = 'A ocurrido un erro al intentar Editar la Ficha'
						msj2='El anbiete esta siendo ocupado en el mismo horario por la ficha numero: '+ocupado
						print('----------->>>>'+msj2)
	else:
		formulario_ficha = agregar_ficha_form(instance = ficha)
	return render(request, 'agregar_ficha.html', locals())
#=================================================#

#==================BORRAR FICHA===================#
def view_eliminar_ficha(request, id_ficha):
	try:
		fc = Persona_ficha.objects.filter(ficha=id_ficha)
		if fc:
			msj = ' ¡Inposible Borrar!. Este PROGRAMA tiene personas asociadas'
			print(">>>>  ¡Inposible Borrar!. Este PROGRAMA tiene personas asociadas")
		else:
			ficha = Ficha.objects.get(id = id_ficha)
			ficha.delete()
	except:
		msj2="Upps a ocurrido u inconveniente. vuelve a intentarlo"

	return redirect('url_lista_fichas')
#=================================================#

#=====================PERMISO=====================#
def view_lista_permisos(request):
	permiso = Permiso.objects.all()
	return render(request, 'lista_permisos.html', locals())
#=================================================#

#=====================PERSONA=====================#

#========== REGISTRAR ADMINISTRADOR ==============#
def view_agregar_administrador(request):
	
	msj = 'Agregar Administrador'

	if request.method == 'POST':
		formulario = agregar_persona_form(request.POST, request.FILES)
		formulario2= agregar_user_form(request.POST, request.FILES)
		formulario4 = elegir_rol_form(request.POST, request.FILES)
		if formulario.is_valid() and formulario2.is_valid() and formulario4.is_valid():

			nom =formulario.cleaned_data['nombres']
			ape =formulario.cleaned_data['apellidos']
			tel =formulario.cleaned_data['telefono']

			documento =formulario.cleaned_data['documentoIdentidad']
			email =formulario2.cleaned_data['email']
			username = str(email)

			rol = formulario4.cleaned_data['rol']

			u = User.objects.create_user(username=username,password=documento, email=email)
			persona = Persona.objects.create(documentoIdentidad=documento,nombres=nom,apellidos=ape,telefono=tel,usuario=u)
			vigilante = Rol_persona.objects.create(rol=rol, persona=persona)
			
			u.save()
			persona.save()
			vigilante.save()
			
			return redirect ('url_index')
	else:
		formulario=agregar_persona_form()
		formulario2= agregar_user_form()
		formulario4 = elegir_rol_form()

	return render(request, 'agregar_administrador.html', locals())
#=================================================#

#==============LISTA INSTRUCTORES=================#
def view_lista_instructores(request):

	try:
		r = Rol.objects.get(rol='Instructor')
		rp = Rol_persona.objects.filter(rol=r)
		
		if not rp:
			msj = 'Lista Vacia. Click aqui para agregar'

	except:
		msj2 = 'Debes agregar El rol Aprendiz'

	return render (request, 'lista_instructores.html', locals())
#=================================================#

#=================REGISTRAR INSTRUCTOR============#
def view_agregar_instructor(request):

	msj = 'Agregar Instructor'

	if request.method == 'POST':
		formulario = agregar_persona_form(request.POST, request.FILES)
		formulario2= agregar_user_form(request.POST, request.FILES)
		formulario4 = elegir_rol_form(request.POST, request.FILES)
		if formulario.is_valid() and formulario2.is_valid() and formulario4.is_valid():

			nom =formulario.cleaned_data['nombres']
			ape =formulario.cleaned_data['apellidos']
			tel =formulario.cleaned_data['telefono']

			documento =formulario.cleaned_data['documentoIdentidad']
			email =formulario2.cleaned_data['email']
			username = str(email)

			rol = formulario4.cleaned_data['rol']

			u = User.objects.create_user(username=username,password=documento, email=email)
			persona = Persona.objects.create(documentoIdentidad=documento,nombres=nom,apellidos=ape,telefono=tel,usuario=u)
			vigilante = Rol_persona.objects.create(rol=rol, persona=persona)
			
			u.save()
			persona.save()
			vigilante.save()
			
			return redirect ('url_lista_instructores')
	else:
		formulario=agregar_persona_form()
		formulario2= agregar_user_form()
		formulario4 = elegir_rol_form()

	return render(request, 'agregar_instructor.html', locals())
#=================================================#

#==============EDITAR_INSTRUCTOR==================#
def view_editar_instructor(request, id_instructor):

	msj = 'Editar Instructor'
	apr = Persona.objects.get(id=id_instructor)
	#fic = Persona_ficha.objects.get(persona=apr)
	if request.method == 'POST':
		formulario = agregar_persona_form(request.POST, request.FILES, instance=apr)
		formulario2 = editar_user_form(request.POST, request.FILES, instance=apr.usuario)
		#formulario3 = elegir_ficha_form(request.POST, request.FILES, instance=fic)
		if formulario.is_valid() and formulario2.is_valid():# and formulario3.is_valid():

			apr = formulario.save()
			usu = formulario2.save()
			#fic = formulario3.save()

			return redirect ('url_lista_instructores')
	else:
		formulario = agregar_persona_form(instance=apr)
		formulario2 = editar_user_form(instance=apr.usuario)
		#formulario3 = elegir_ficha_form(instance=fic)
	return render (request, 'agregar_instructor.html', locals())
#=================================================#

#==============ELIMINAR_INSTRUCTOR================#
def view_eliminar_instructor(request, id_instructor):
	try:
		instructor = Persona.objects.get(id=id_instructor)
		id_user = instructor.usuario.id
		user = User.objects.get(id=id_user)
		instructor.delete()
		user.delete()
	except:
		msj2="Upps a ocurrido u inconveniente. vuelve a intentarlo"

	return redirect ('url_lista_instructores')
#=================================================#

#==============LISTA_VIGILANTES==================#
def view_lista_vigilantes(request):
	try:
		vg= Rol.objects.get(rol='Vigilante')
		vig=Rol_persona.objects.filter(rol=vg)
		if not vig:
			msj = 'Lista Vacia. Click aqui para agregar'
	except:
		msj2 = 'Debes agregar El rol Vigilante'

	return render(request, 'lista_vigilantes.html', locals())
#================================================#

#===========REGISTRAR VIGILANTE===================#
def view_agregar_vigilante(request):

	if request.method == 'POST':
		formulario = agregar_persona_form(request.POST, request.FILES)
		formulario2= agregar_user_vigilante_form(request.POST, request.FILES)
		formulario4 = elegir_rol_form(request.POST, request.FILES)
		if formulario.is_valid() and formulario2.is_valid() and formulario4.is_valid():

			nom =formulario.cleaned_data['nombres']
			ape =formulario.cleaned_data['apellidos']
			tel =formulario.cleaned_data['telefono']

			documento =formulario.cleaned_data['documentoIdentidad']
			email =formulario2.cleaned_data['username']

			rol = formulario4.cleaned_data['rol']

			u = User.objects.create_user(username=email,password=documento, email=email)
			persona = Persona.objects.create(documentoIdentidad=documento,nombres=nom,apellidos=ape,telefono=tel,usuario=u)
			vigilante = Rol_persona.objects.create(persona=persona, rol=rol)
			
			u.save()
			persona.save()
			vigilante.save()
			
			return redirect ('url_lista_vigilantes')

	else:
		formulario=agregar_persona_form()
		formulario2= agregar_user_vigilante_form()
		formulario4 = elegir_rol_form()
	return render(request, 'agregar_instructor.html', locals())	
#=================================================#

#==============EDITAR_VIGILANTE====================#
def view_editar_vigilante(request, id_vigilante):
	vgt= Persona.objects.get(id= id_vigilante)
	if request.method== 'POST':
		formulario= agregar_persona_form(request.POST, request.FILES, instance= vgt)
		formulario2= editar_user_form(request.POST, request.FILES, instance= vgt.usuario)
		if formulario.is_valid() and formulario2.is_valid():

			vgt= formulario.save()
		
			vgt.usuario.set_password(formulario2.cleaned_data['contraseña'])

			vgt= formulario2.save()

			return redirect('url_lista_vigilantes')
	else:
		formulario = agregar_persona_form(instance=vgt)
		formulario2 = editar_user_form(instance=vgt.usuario)
	return render (request, 'agregar_instructor.html', locals())
#=================================================#

#==============ELIMINAR_VIGILANTE====================#
def view_eliminar_vigilante(request, id_vigilante):
	try:
		vigilante = Persona.objects.get(id=id_vigilante)
		id_user = vigilante.usuario.id
		user = User.objects.get(id=id_user)
		vigilante.delete()
		user.delete()
	except:
		msj2="Upps a ocurrido u inconveniente. vuelve a intentarlo"

	return redirect ('url_lista_vigilantes')
#=================================================#

#==============LISTA_APRENDICES===================#
def view_lista_aprendices(request):

	try:
		r = Rol.objects.get(rol='Aprendiz')
		rp = Rol_persona.objects.filter(rol=r)
		
		if not rp:
			msj = 'Lista Vacia. Click aqui para agregar'

	except:
		msj2 = 'Debes agregar El rol Aprendiz'

	return render (request, 'lista_aprendices.html', locals())
#=================================================#

#==============AGREGAR_APRENDIZ===================#
def view_agregar_aprendiz(request):

	if request.method == 'POST':
		formulario = agregar_persona_form(request.POST, request.FILES)
		formulario2 = agregar_user_form(request.POST, request.FILES)
		formulario3 = elegir_ficha_form(request.POST, request.FILES)
		formulario4 = elegir_rol_form(request.POST, request.FILES)
		if formulario.is_valid() and formulario2.is_valid() and formulario3.is_valid() and formulario4.is_valid():

			nom =formulario.cleaned_data['nombres']
			ape =formulario.cleaned_data['apellidos']
			tel =formulario.cleaned_data['telefono']

			documento =formulario.cleaned_data['documentoIdentidad']
			email =formulario2.cleaned_data['email']
			username = str(email)

			fic = formulario3.cleaned_data['ficha']
			rol = formulario4.cleaned_data['rol']

			#Objetos
			u = User.objects.create_user(username=username,password=documento, email=email)
			persona = Persona.objects.create(documentoIdentidad=documento,nombres=nom,apellidos=ape,telefono=tel,usuario=u)
			fich_per = Persona_ficha.objects.create(persona=persona, ficha=fic)
			aprendiz = Rol_persona.objects.create(persona=persona, rol=rol)
			
			#Guardar Objetos
			u.save()
			persona.save()
			fich_per.save()
			aprendiz.save()

			return redirect ('url_lista_aprendices')

	else:
		formulario = agregar_persona_form()
		formulario2 = agregar_user_form()
		formulario3 = elegir_ficha_form()
		formulario4 = elegir_rol_form()

	return render (request, 'agregar_aprendiz.html', locals())
#=================================================#

#==============EDITAR_APRENDIZ====================#
def view_editar_aprendiz(request, id_aprendiz):
	apr = Persona.objects.get(id=id_aprendiz)
	fic = Persona_ficha.objects.get(persona=apr)

	if request.method == 'POST':
		formulario = agregar_persona_form(request.POST, request.FILES, instance=apr)
		formulario2 = editar_user_form(request.POST, request.FILES, instance=apr.usuario)
		formulario3 = elegir_ficha_form(request.POST, request.FILES, instance=fic)
		if formulario.is_valid() and formulario2.is_valid() and formulario3.is_valid():

			apr = formulario.save()
			usu = formulario2.save()
			fic = formulario3.save()

			return redirect ('url_lista_aprendices')
	else:
		formulario = agregar_persona_form(instance=apr)
		formulario2 = editar_user_form(instance=apr.usuario)
		formulario3 = elegir_ficha_form(instance=fic)
	return render (request, 'agregar_aprendiz.html', locals())
#=================================================#

#==============ELIMINAR_APRENDIZ==================#
def view_eliminar_aprendiz(request, id_aprendiz):
	try:
		aprendiz = Persona.objects.get(id=id_aprendiz)
		id_user = aprendiz.usuario.id
		user = User.objects.get(id=id_user)
		aprendiz.delete()
		user.delete()
	except:
		msj2="Upps a ocurrido u inconveniente. vuelve a intentarlo"
	return redirect ('url_lista_aprendices')
#=================================================#

#=========VIEW_AGREGAR_DESDE_EXCEL================#
def view_agregar_aprendiz_excel(request):
	fichas = Ficha.objects.all()
	if request.method == 'POST':
		form = cargar_excel_form(request.POST, request.FILES)
		if form.is_valid():
			excel = form.cleaned_data['excel']
			return redirect('url_lista_aprendices')#'url_registros_excel')
	else:
		form = cargar_excel_form()
	return render(request, 'agregar_aprendices_excel.html', locals())
#=================================================#

#=============AGREGAR_DESDE_EXCEL=================#
def cargar_excel(request):
	rl = Rol.objects.get(rol='Aprendiz')
	FILE_PATH = 'media/registros.xlsx' #define el achivo excel en una variable
	SHEET = 'Hoja1'                  #define la hoja de excel en una variable

	wb = load_workbook(FILE_PATH, read_only=True)# cargo el archivo en una variable
	sheet = wb[SHEET]                            # cargo la hoja en una variable
	
	for row in sheet.iter_rows(min_row=2):						
		documento = row[0].value
		nom = row[1].value
		ape = row[2].value
		email = row[3].value
		#ape2 = row[3].value
		tel = row[5].value

		u = User.objects.create_user(username=email,password=documento, email=email)
		u.save()
		
		persona = Persona.objects.create(documentoIdentidad=documento,nombres=nom,apellidos=ape,telefono=tel,usuario=u)
		persona.save()


		aprendiz = Rol_persona.objects.create(rol=rl ,persona=persona)
		aprendiz.save()
	lista = sheet.iter_rows(min_row=2)

	return render(request, 'registros_excel.html', locals())
#=================================================#

#====================PROGRAMA=====================#
def view_lista_programas(request):
	#LISTAR
	programa = Programa.objects.all()
	if not programa:
			msj = 'Lista Vacia. Click aqui para agregar'
	return render(request, 'lista_programas.html', locals())
#=================================================#

#===============AGREGAR PROGRAMA==================#
def view_agregar_programa(request):
	if request.method == 'POST':
		formulario = agregar_programa_form(request.POST, request.FILES)
		if formulario.is_valid():
			pro = formulario.save(commit=False)
			pro.save()
			return redirect('url_lista_programas')
	else:
		formulario = agregar_programa_form()

	return render (request, 'agregar_programa.html',locals())
#=================================================#

#===================EDITAR PROGRAMA===============#	
def view_editar_programa(request, id_programa):
	pro = Programa.objects.get(id=id_programa)
	if request.method == "POST":
		formulario = agregar_programa_form(request.POST, request.FILES, instance=pro)
		if formulario.is_valid():
			pro = formulario.save()
			return redirect('url_lista_programas')
	else:
		formulario = agregar_programa_form(instance= pro)
	return render(request, 'agregar_programa.html', locals())
#=================================================#

#===================ELIMINAR PROGRAMA=============#
def view_eliminar_programa(request, id_programa):
	try:	
		pr = Ficha.objects.filter(programa=id_programa)
		if pr:
			msj = ' ¡Inposible Borrar!. Este PROGRAMA tiene personas asociadas'
		else:
			pro = Programa.objects.get(id=id_programa)
			pro.delete()
	except:
		msj2="Upps a ocurrido u inconveniente. vuelve a intentarlo"

	return redirect('url_lista_programas')	
#=================================================#

#=====================USUARIO=====================#
def view_usuario(request):
	usuario = User.objects.all()
	return render(request, 'usuario/index.html', locals())
#=================================================#

#===================VER_ROL=======================#
def view_lista_roles(request):
	rol = Rol.objects.all()
	if not rol:
			msj = 'Lista Vacia. Click aqui para agregar'
	return render(request,'lista_roles.html',locals())
#=================================================#

#===================AGREGAR ROL===================#
def view_agregar_rol(request):
	if request.method== 'POST':
		formulario= agregar_rol_form(request.POST,request.FILES)
		if formulario.is_valid():
			rol= formulario.save(commit= False)
			rol.save()
			formulario.save_m2m()
			return redirect('url_lista_roles')

	else:
		formulario = agregar_rol_form()
	return render(request,'agregar_rol.html',locals())
#=================================================#

#===================EDITAR_ROL====================#
def view_editar_rol(request, id_rol):
	rol= Rol.objects.get(id=id_rol)
	if request.method== 'POST':
		formulario= agregar_rol_form(request.POST, request.FILES, instance=rol)
		if formulario.is_valid():
			rol= formulario.save()
			return redirect('url_lista_roles')
	else:
		formulario= agregar_rol_form(instance= rol)
	return render(request, 'agregar_rol.html',locals())
#=================================================#

#===================ELIMINAR_ROL==================#
def view_eliminar_rol(request, id_rol):
	try:
		rp = Rol_persona.objects.filter(rol=id_rol)
		if rp:
			msj = ' ¡Inposible Borrar! Este Rol tiene personas asociadas'
			print(">>>> "+msj)
		else:
			rol= Rol.objects.get(id= id_rol)
			rol.delete()
	except:
		msj="Upps a ocurrido u inconveniente. vuelve a intentarlo"

	return redirect('url_lista_roles')
#==================== LOGIN ======================#
#@login_required  DECORADORES
def view_login(request):
	error = "Usted no esta autorizado para ingresar desde este terminal. Por favor dirijase a su Teléfono."
	comprobacion_rol = ""
	info_enviada = False
	if request.user.is_authenticated:
		return redirect('url_index')
	else:
		loginf = login_form()
		if request.method == "POST":
			loginf = login_form(request.POST)
			if loginf.is_valid():
				usuario = loginf.cleaned_data['usuario']
				clave = loginf.cleaned_data["contraseña"]
				user = authenticate(username = usuario, password = clave)
				try:
					user_temp = User.objects.get(username = usuario)
					user_id_temp = user_temp.id
					per_id= Persona.objects.get(usuario_id = user_id_temp)
					per_temp = per_id.id
					rol_temp = Rol_persona.objects.get(persona_id = per_temp)
					comprobacion_rol = str(rol_temp.rol)

					if comprobacion_rol == "ADMINISTRADOR" or comprobacion_rol == "VIGILANTE":
						if user is not None and user.is_active:
							login(request, user)
							usuarioActivo = True
							return redirect('url_index')
						else:
							msj = "usuario o clave incorrecto"
					else:
						info_enviada = True
						loginf = login_form()
						return render(request, "login.html", locals())
				except:
					msj = "El Usuario no existe"
		loginf = login_form()
		return render(request, "login.html", locals())
#=================================================#


#==================== LOGOUT =====================#
def view_logout(request):
	logout(request)
	return redirect('url_index')
#=================================================#

#=================================================#
